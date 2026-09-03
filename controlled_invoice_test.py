"""Explicitly authorized Group B test. No CLI and no automatic retries."""
import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal

import openpyxl
import requests
from filelock import FileLock

import core
from partner_export import prepare_partner_export

PAYOUTS = ('7700379513', '7710027297', '7712804241')
CONTACT = 'a9f15779-2ab7-4905-9c04-63a6bf979f6d'


def require(condition, message):
    if not condition:
        raise ValueError(message)


def prepare(workbook):
    master = core.load_master_data()
    rows = master[(master['Gruppe'] == 'Gruppe B') & (master['Art'] == 'Bestellung') & (master['Erlös_Brutto'] > 0)]
    require(len(rows) == 37 and rows['Partner'].value_counts().to_dict() == {'MH': 25, 'NB': 12}, 'Positionsumfang abweichend')
    require(set(rows['Auszahlung Nr.']) == set(PAYOUTS), 'Payoutumfang abweichend')
    model = prepare_partner_export(rows, statement_type='group_b_evelyn')
    totals = model['totals']['Rechnung']
    require(all(totals[k] == Decimal(v) for k, v in {'net_after': '3496.06', 'tax': '664.25', 'gross': '4160.31'}.items()), 'Quellsummen abweichend')
    book = openpyxl.load_workbook(workbook, data_only=True)
    sheet = book['Rechnung']
    actual = Counter((sheet.cell(r, 3).value, sheet.cell(r, 4).value, Decimal(str(sheet.cell(r, 7).value))) for r in range(15, 52))
    expected = Counter((item['article'], item['extra'], item['net']) for item in model['Rechnung'])
    require(actual == expected, 'Excel und Originaldaten weichen ab')
    require(all(sheet.cell(r, 5).value == 1 and sheet.cell(r, 6).value == 'Stück' and sheet.cell(r, 8).value == .005 and sheet.cell(r, 9).value == .19 for r in range(15, 52)), 'Excel-Positionsparameter abweichend')
    book.close()
    now = datetime.now(timezone.utc).isoformat(timespec='milliseconds')
    payload = {
        'voucherDate': now, 'address': {'contactId': CONTACT},
        'lineItems': [{'type': 'custom', 'name': x['article'], 'description': x['extra'], 'quantity': 1, 'unitName': 'Stück',
                       'unitPrice': {'currency': 'EUR', 'netAmount': float(x['net']), 'taxRatePercentage': 19},
                       'discountPercentage': .5} for x in model['Rechnung']],
        'totalPrice': {'currency': 'EUR'}, 'taxConditions': {'taxType': 'net'},
        'shippingConditions': {'shippingDate': now, 'shippingType': 'service'},
        'remark': core.invoice_payout_remark(rows['Auszahlung Nr.']),
    }
    return master, payload, hashlib.sha256(workbook.read_bytes()).hexdigest()


def reserve(master, payload, workbook_hash):
    with core.ledger() as db:
        db.execute('BEGIN IMMEDIATE')
        for payout in PAYOUTS:
            row = db.execute('SELECT * FROM payouts WHERE id=?', (payout,)).fetchone()
            require(row is not None and not row['attempt'] and not row['invoice_id'], 'Bereits reserviert/erstellt; kein weiterer Versuch erlaubt')
        for payout in PAYOUTS:
            block = master[master['Auszahlung Nr.'] == payout]
            db.execute("UPDATE payouts SET attempt='pending', status='Einmaliger Gesamtentwurf reserviert', fingerprint=?, snapshot=? WHERE id=?",
                       (core.payout_fingerprint(block), json.dumps(payload, ensure_ascii=False), payout))
            core.audit(db, payout, 'Explizit freigegebener EINMALIGER Gruppe-B-Test; dauerhaft reserviert vor POST; Excel SHA256=' + workbook_hash)
        db.commit()


def create_once(api_key, workbook, http=None):
    require(bool(api_key), 'API-Schlüssel fehlt')
    session = http or requests.Session()
    if http is None:
        session.mount('https://', requests.adapters.HTTPAdapter(max_retries=0))
    headers = {'Authorization': 'Bearer ' + api_key, 'Accept': 'application/json'}
    with FileLock(core.PAYOUTS_DB_PATH + '.lock'), FileLock(core.ORDERS_DB_PATH + '.lock'):
        master, payload, workbook_hash = prepare(workbook)
        contact = session.get(core.API_URL + '/contacts/' + CONTACT, headers=headers, timeout=20)
        require(contact.status_code == 200, 'Kontakt-Lesezugriff fehlgeschlagen; kein POST')
        require(str(contact.json().get('roles', {}).get('customer', {}).get('number')) == '16335', 'Kontakt abweichend; kein POST')
        reserve(master, payload, workbook_hash)
    # This is the ONLY creation call. A persisted reservation blocks every rerun.
    try:
        response = session.post(core.API_URL + '/invoices', params={'finalize': 'false'}, headers=headers,
                                json=payload, timeout=(10, 30), allow_redirects=False)
        require(response.status_code in (200, 201), 'HTTP ' + str(response.status_code))
        invoice_id = response.json().get('id')
        require(isinstance(invoice_id, str) and bool(invoice_id), 'Erstellungs-ID fehlt')
    except Exception:
        with core.ledger() as db:
            for payout in PAYOUTS:
                db.execute("UPDATE payouts SET attempt='unknown', status='Prüfung erforderlich' WHERE id=?", (payout,))
                core.audit(db, payout, 'Ergebnis unklar/fehlerhaft; dauerhaft gesperrt; nur lesend abgleichen')
            db.commit()
        raise RuntimeError('Erstellungsantwort nicht eindeutig erfolgreich. Kein weiterer POST erlaubt; nur lesend abgleichen.') from None
    with core.ledger() as db:
        for payout in PAYOUTS:
            db.execute("UPDATE payouts SET attempt='created', invoice_id=?, status='Lexoffice-Entwurf erstellt' WHERE id=?", (invoice_id, payout))
            core.audit(db, payout, 'Gesamtentwurf erstellt: ' + invoice_id)
        db.commit()
    return invoice_id, payload


def verify(saved, payload):
    errors = []
    for key in ('name', 'description', 'quantity', 'unitName', 'discountPercentage'):
        if [i.get(key) for i in saved.get('lineItems', [])] != [i[key] for i in payload['lineItems']]:
            errors.append(key)
    for key in ('netAmount', 'taxRatePercentage', 'currency'):
        if [i.get('unitPrice', {}).get(key) for i in saved.get('lineItems', [])] != [i['unitPrice'][key] for i in payload['lineItems']]:
            errors.append('unitPrice.' + key)
    if len(saved.get('lineItems', [])) != 37:
        errors.append('Positionsanzahl')
    for key, expected in {'totalNetAmount': '3496.06', 'totalTaxAmount': '664.25', 'totalGrossAmount': '4160.31'}.items():
        if Decimal(str(saved.get('totalPrice', {}).get(key))) != Decimal(expected):
            errors.append(key)
    if saved.get('remark') != payload['remark']:
        errors.append('Freitext')
    if saved.get('voucherStatus') != 'draft':
        errors.append('Entwurfsstatus')
    if saved.get('address', {}).get('contactId') != CONTACT:
        errors.append('Kontakt')
    return errors
