import hashlib
import io
import json
import os
import shutil
import tempfile
import unittest
from datetime import datetime
from decimal import Decimal
from fractions import Fraction
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook  # Independent read-only verification.

import core
from partner_export import export_partner_excel, prepare_partner_export, report_date, calculate_sheet
from test_recovery import payout


def reference_cents(value):
    """Independent integer/rational half-away-from-zero rounding oracle."""
    value = Fraction(value) * 100
    absolute = abs(value)
    whole = (2 * absolute.numerator + absolute.denominator) // (2 * absolute.denominator)
    return (-whole if value < 0 else whole)


def check_workbook(case, blob, rows, rate, recipient):
    book = load_workbook(io.BytesIO(blob), data_only=True)
    formula_book = load_workbook(io.BytesIO(blob), data_only=False)
    case.assertEqual(book.sheetnames, ['Rechnung', 'Gutschriften'])
    payout_ids = set(rows['Auszahlung Nr.'])
    for name, kind in [('Rechnung', 'Bestellung'), ('Gutschriften', 'Erstattung')]:
        expected = rows[rows.Art == kind]
        sheet = book[name]
        last = 14 + max(1, len(expected))
        summary = last + 2
        case.assertEqual(sheet.max_column, 11)
        case.assertEqual(sheet.freeze_panes, 'A15')
        case.assertEqual(sheet['E4'].value, recipient)
        case.assertEqual(sheet['G4'].value, float(rate))
        case.assertEqual(sheet['I4'].value, .19)
        case.assertEqual(set(sheet['E6'].value.split(', ')), payout_ids)
        case.assertEqual(sheet['A6'].value, 'Rechnungsadresse noch nicht hinterlegt')
        text = '\n'.join(str(cell.value) for row in sheet for cell in row if cell.value is not None)
        case.assertNotIn('provision', text.lower())
        case.assertIn('Freitext auf der Rechnung', text)
        net_before = net_after = ebay = previous_tax = gross_sum = 0
        for number, (_, original) in enumerate(expected.iterrows(), 15):
            case.assertEqual(sheet[f'E{number}'].value, 1)
            case.assertEqual(sheet[f'F{number}'].value, 'Stück')
            case.assertEqual(sheet[f'G{number}'].value, original['eBay_Netto'])
            case.assertEqual(sheet[f'C{number}'].value, original['Angebotstitel'])
            case.assertEqual(sheet[f'D{number}'].value,
                             'eBay-Bestellnummer: '+original['Bestellnummer']+'\nSKU: '+original['SKU'])
            case.assertIsInstance(sheet[f'A{number}'].value, datetime)
            case.assertTrue(sheet[f'C{number}'].alignment.wrap_text)
            case.assertEqual(sheet[f'G{number}'].alignment.horizontal, 'right')
            case.assertEqual(sheet[f'E{number}'].alignment.horizontal, 'center')
            case.assertNotEqual(sheet[f'C{number}'].fill.fgColor.rgb, sheet[f'A{number}'].fill.fgColor.rgb)
            net = Fraction(str(original['eBay_Netto']))
            after = reference_cents(net * (1-Fraction(rate)))
            net_before += reference_cents(net)
            net_after += after
            tax = reference_cents(Fraction(net_after,100)*Fraction(19,100))
            gross = after + tax - previous_tax
            case.assertEqual(reference_cents(str(sheet[f'J{number}'].value)), gross)
            gross_sum += gross
            previous_tax = tax
            ebay += reference_cents(str(sheet[f'K{number}'].value))
            formula = formula_book[name][f'J{number}'].value
            helper=summary+11+number-15
            case.assertEqual(formula_book[name][f'G{helper}'].value,f'=E{number}*G{number}')
            case.assertEqual(formula_book[name][f'H{helper}'].value,f'=ROUND(G{helper}*(1-H{number}),2)')
            case.assertTrue(sheet.row_dimensions[helper].hidden)
            case.assertNotIn('K', formula)  # The eBay control never drives the invoice.
        expected_totals = [net_before, net_before-net_after, net_after, previous_tax,
                           net_after+previous_tax, ebay, ebay-net_after-previous_tax]
        for offset, expected_cents in enumerate(expected_totals):
            case.assertEqual(reference_cents(str(sheet[f'K{summary+offset}'].value)), expected_cents)
        case.assertEqual(gross_sum, expected_totals[4])
    return book


class PartnerExportTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.paths = patch.multiple(core, PAYOUTS_DB_PATH=str(Path(self.temp.name)/'Master_Payouts.csv'),
                                    ORDERS_DB_PATH=str(Path(self.temp.name)/'Master_Orders.csv'))
        self.paths.start()
        self.addCleanup(self.paths.stop)
        self.http = patch('requests.sessions.Session.request', side_effect=AssertionError('HTTP forbidden'))
        self.http.start()
        self.addCleanup(self.http.stop)

    def seed(self, sku='NB / TEST', refund=False):
        order = payout(sku=sku, title='Vollständiger Produkttitel ' * 6)
        order['Verkauft am'] = '25-Aug-26'
        order['Anzahl'] = '7'
        core.import_reports([order], core.ORDERS_DB_PATH, 'orders')
        sale = payout(sku='WRONG', title='Abgeschnittener Payout-Titel', amount='62,99')
        sale['Auszahlungsdatum'] = '2. Sep 2026'
        sale['Transaktionsbetrag (inkl. Kosten)'] = '69,99'
        frames = [sale]
        if refund:
            credit = payout(transaction='refund1', amount='-9,90', kind='Rückerstattung')
            credit['Auszahlungsdatum'] = '2. Sep 2026'
            credit['Transaktionsbetrag (inkl. Kosten)'] = '-9,90'
            frames.append(credit)
        core.import_reports(frames, core.PAYOUTS_DB_PATH, 'payout')
        return core.load_master_data()

    def test_schema_sources_and_negative_refund(self):
        master = self.seed(refund=True)
        before = master.copy(deep=True)
        book = check_workbook(self, export_partner_excel(master), master, Decimal('.035'), 'Patrick')
        self.assertEqual(book['Rechnung']['K15'].value, 69.99)
        self.assertEqual(book['Rechnung']['J15'].value, 60.79)
        self.assertEqual(book['Rechnung']['A15'].number_format, 'dd"."mm"."yyyy')
        self.assertGreaterEqual(book['Rechnung'].row_dimensions[15].height, 76)
        self.assertIn('€', book['Rechnung']['G15'].number_format)
        core.pd.testing.assert_frame_equal(master, before)

    def test_all_groups_and_automatic_rates(self):
        for partner in ('PP', 'BA', 'MK', '001', 'NB', 'MH12', 'OTHER'):
            with self.subTest(partner=partner), tempfile.TemporaryDirectory() as folder, patch.object(core, 'known_group_b_partners', return_value={'NB', 'OTHER'}):
                with patch.multiple(core, PAYOUTS_DB_PATH=str(Path(folder)/'Master_Payouts.csv'),
                                    ORDERS_DB_PATH=str(Path(folder)/'Master_Orders.csv')):
                    master = self.seed(sku=partner+' / TEST')
                    rate = Decimal('.005') if partner in ('PP','BA','MK','001') else Decimal('.035')
                    book = check_workbook(self, export_partner_excel(master), master, rate,
                                          'Evelyn' if rate == Decimal('.005') else 'Patrick')
                    self.assertIn('Keine Erstattungen', book['Gutschriften']['C15'].value)
                    if partner.startswith('MH'):
                        self.assertEqual(book['Rechnung']['A4'].value, 'MH')
                    if master.iloc[0].Gruppe == 'Gruppe B':
                        check_workbook(self, export_partner_excel(master,statement_type='group_b_evelyn'),
                                       master,Decimal('.005'),'Evelyn')

    def test_rounding_regression_and_column_tax(self):
        # NB example: old independent 392.50 * .965 = 378.7625, displayed 378.76.
        items = [{'net':Decimal('329.83'),'ebay':Decimal('392.50')}]
        totals = calculate_sheet(items,Decimal('.035'))
        self.assertEqual(totals['net_after'],Decimal('318.29'))
        self.assertEqual(totals['gross'],Decimal('378.77'))
        # Two individually rounded VAT values would produce .14, column VAT .13.
        items = [{'net':Decimal('.33'),'ebay':Decimal('.39')} for _ in range(2)]
        totals = calculate_sheet(items,Decimal('.005'))
        self.assertEqual(totals['gross'],Decimal('.79'))
        self.assertEqual(sum(item['gross'] for item in items),Decimal('.79'))
        for sign in (1,-1):
            items=[{'net':Decimal('1.00')*sign,'ebay':Decimal('1.19')*sign}]
            self.assertEqual(calculate_sheet(items,Decimal('.005'))['net_after'],Decimal('1.00')*sign)

    def test_missing_original_control_amount_blocks_export(self):
        master = self.seed()
        raw = core.read_master(core.PAYOUTS_DB_PATH).drop(columns=['Transaktionsbetrag (inkl. Kosten)'])
        with self.assertRaisesRegex(ValueError,'Bruttobetrag fehlt'):
            export_partner_excel(master,payouts=raw)

    def test_refund_only_and_dates(self):
        master = self.seed(refund=True)
        check_workbook(self,export_partner_excel(master[master.Art=='Erstattung']),
                       master[master.Art=='Erstattung'],Decimal('.035'),'Patrick')
        for text in ('2. Sep 2026','02-Sep-26','02.09.2026','2026-09-02'):
            self.assertEqual(report_date(text),datetime(2026,9,2))
        with self.assertRaises(ValueError):
            report_date('31. Feb 2026')

    def test_recipient_master_data_and_invalid_scope(self):
        master = self.seed(sku='BA / TEST')
        with self.assertRaises(ValueError):
            export_partner_excel(master,statement_type='group_b_evelyn')
        with self.assertRaises(ValueError):
            export_partner_excel(master,statement_type='manual_rate')
        from partner_export import recipient_details
        config=json.loads(Path('billing_recipients.json').read_text(encoding='utf-8'))
        config['recipients']['evelyn']['address']['street']='TEST-STRASSE (synthetisch)'
        path=Path(self.temp.name)/'recipients.json'
        path.write_text(json.dumps(config),encoding='utf-8')
        with patch.dict(os.environ,{'PAYMENT_RECIPIENTS_PATH':str(path)}):
            self.assertEqual(recipient_details('evelyn')[1],'TEST-STRASSE (synthetisch)')


@unittest.skipUnless(os.environ.get('EBAY_REAL_MASTER_DIR'),'Set EBAY_REAL_MASTER_DIR for original imported data')
class RealPartnerExportTests(unittest.TestCase):
    def test_all_original_positions_and_four_outputs(self):
        source=Path(os.environ['EBAY_REAL_MASTER_DIR'])
        filenames=['Master_Payouts.csv','Master_Orders.csv']
        before={name:hashlib.sha256((source/name).read_bytes()).hexdigest() for name in filenames}
        with tempfile.TemporaryDirectory() as folder:
            # This isolated offline fixture intentionally starts with a fresh ledger.
            with patch.object(core, 'PAYOUTS_DB_PATH', str(Path(folder)/filenames[0])):
                with core.ledger():
                    pass
            for name in filenames:
                shutil.copyfile(source/name,Path(folder)/name)
            with patch.multiple(core,PAYOUTS_DB_PATH=str(Path(folder)/filenames[0]),ORDERS_DB_PATH=str(Path(folder)/filenames[1])), \
                 patch('requests.sessions.Session.request',side_effect=AssertionError('HTTP forbidden')):
                master=core.load_master_data()
                self.assertEqual(len(master),50)
                self.assertEqual(len(core.read_master(core.ORDERS_DB_PATH)),180)
                self.assertAlmostEqual(master['Erlös_Brutto'].sum(),4427.83)
                cases=[('BA',master[master.Partner=='BA'],'partner',Decimal('.005'),'Evelyn',1,0,'62.68','0.00'),
                       ('NB',master[master.Partner=='NB'],'partner',Decimal('.035'),'Patrick',12,2,'2007.68','-347.11'),
                       ('MH',master[master.Partner=='MH'],'partner',Decimal('.035'),'Patrick',25,3,'2027.21','-92.58'),
                       ('Gruppe_B_Evelyn',master[master.Gruppe=='Gruppe B'],'group_b_evelyn',Decimal('.005'),'Evelyn',37,5,'4160.31','-453.38')]
                verification={}
                for filename,rows,kind,rate,recipient,sales,refunds,gross,credit in cases:
                    with self.subTest(file=filename):
                        self.assertEqual((rows.Art=='Bestellung').sum(),sales)
                        self.assertEqual((rows.Art=='Erstattung').sum(),refunds)
                        self.assertFalse(rows.duplicated().any())
                        model=prepare_partner_export(rows,statement_type=kind)
                        self.assertEqual(model['totals']['Rechnung']['gross'],Decimal(gross))
                        self.assertEqual(model['totals']['Gutschriften']['gross'],Decimal(credit))
                        blob=export_partner_excel(rows,statement_type=kind)
                        book=check_workbook(self,blob,rows,rate,recipient)
                        output=os.environ.get('PARTNER_TEST_OUTPUT_DIR')
                        if output:
                            Path(output).mkdir(parents=True,exist_ok=True)
                            (Path(output)/(filename+'.xlsx')).write_bytes(blob)
                            formulas=load_workbook(io.BytesIO(blob),data_only=False)
                            verification[filename+'.xlsx']={sheet.title:{
                                'lastRow':min(r for r,d in sheet.row_dimensions.items() if d.hidden)-3,
                                'formulas':{cell.coordinate:book[sheet.title][cell.coordinate].value
                                            for row in sheet for cell in row if cell.data_type=='f'}
                            } for sheet in formulas}
                if os.environ.get('PARTNER_TEST_OUTPUT_DIR'):
                    (Path(os.environ['PARTNER_TEST_OUTPUT_DIR'])/'verification.json').write_text(
                        json.dumps(verification),encoding='utf-8')
                self.assertEqual(set(master[master.Gruppe=='Gruppe B'].Partner),{'NB','MH'})
                from streamlit.testing.v1 import AppTest
                app=AppTest.from_file(str(Path(__file__).with_name('app.py'))).run()
                for pid in sorted(master['Auszahlung Nr.'].unique()):
                    next(widget for widget in app.selectbox if widget.label=='Eine Auszahlung wählen').select(pid).run()
                    self.assertFalse(list(app.exception))
                    self.assertFalse(any('Partnerexport angehalten' in error.value for error in app.error))
        for name in filenames:
            self.assertEqual(hashlib.sha256((source/name).read_bytes()).hexdigest(),before[name])


if __name__=='__main__':
    unittest.main()
